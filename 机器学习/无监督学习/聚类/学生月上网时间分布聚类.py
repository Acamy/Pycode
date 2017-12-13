import numpy as np
import openpyxl
import sklearn.cluster as skc
from sklearn import metrics
import matplotlib.pyplot as plt



wb = openpyxl.load_workbook('TestData.xlsx')
sh = wb.get_sheet_by_name('Sheet1')
mac2id=dict()
onlinetimes=[]
for i in range(2, sh.max_row + 1):
    mac = sh.cell(row=i, column=3).value
    starttime = int(sh.cell(row=i, column=5).value.timestamp())
    endtime = int(sh.cell(row=i, column=6).value.timestamp())
    onlinetime = endtime - starttime
    if mac not in mac2id:
        mac2id[mac] = len(onlinetimes)
        onlinetimes.append((starttime, onlinetime))
    else:
        onlinetimes[mac2id[mac]] = [(starttime, onlinetime)]

real_X=np.array(onlinetimes).reshape((-1,2))
 
X=real_X[:,0:1]
 
db=skc.DBSCAN(eps=0.01,min_samples=20).fit(X)
labels = db.labels_
 
print('Labels:')
print(labels)
raito=len(labels[labels[:] == -1]) / len(labels)
print('Noise raito:',format(raito, '.2%'))
 
n_clusters_ = len(set(labels)) - (1 if -1 in labels else 0)
 
print('Estimated number of clusters: %d' % n_clusters_)
print("Silhouette Coefficient: %0.3f"% metrics.silhouette_score(X, labels))
 
for i in range(n_clusters_):
    print('Cluster ',i,':')
    print(list(X[labels == i].flatten()))
     
plt.hist(X,24)