import numpy as np
from sklearn.cluster import KMeans
import openpyxl
 
def loadData(filePath):
    wb = openpyxl.load_workbook(filePath)
    sh = wb.get_sheet_by_name('Sheet1')
    retData = []
    retCityName = []
    for i in range(5, sh.max_row + 1):
        retCityName.append(sh.cell(row=i, column=1).value)
        retData.append([float(sh.cell(row=i, column=j).value) for j in range(2,sh.max_column + 1)])
    return retData,retCityName
 
     
if __name__ == '__main__':
    data,cityName = loadData('city.xlsx')
    km = KMeans(n_clusters=5)
    label = km.fit_predict(data)
    expenses = np.sum(km.cluster_centers_,axis=1)
    #print(expenses)
    CityCluster = [[],[],[],[],[]]
    for i in range(len(cityName)):
        CityCluster[label[i]].append(cityName[i])
    for i in range(len(CityCluster)):
        print("Expenses:%.2f" % expenses[i])
        print(CityCluster[i])

